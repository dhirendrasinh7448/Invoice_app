from pathlib import Path

import pandas as pd
import glob
import openpyxl
from fpdf import FPDF

filepaths = glob.glob("xlsx_files/*.xlsx")

for index, filepath in enumerate(filepaths):
    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_nr, date = filename.split("-")

    pdf.set_font(family="Times", style="B", size=16)
    pdf.cell(w=50, h=8, txt=f"Invoice_nr {invoice_nr}", ln=1)

    pdf.set_font(family="Times", style="B", size=16)
    pdf.cell(w=50, h=8, txt=f"Date: {date}", ln=1)

    columns = df.columns
    columns = [item.replace("_", " ").title() for item in columns]
    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(80, 80, 80)

    # Set The Header
    pdf.cell(w=30, h=8, txt=columns[0], border=1)
    pdf.cell(w=50, h=8, txt=columns[1], border=1)
    pdf.cell(w=30, h=8, txt=columns[2], border=1)
    pdf.cell(w=30, h=8, txt=columns[3], border=1)
    pdf.cell(w=30, h=8, txt=columns[4], border=1, ln=1)

    # Products Rows
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt= str(row["product_id"]), border=1)
        pdf.cell(w=50, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    # Total Sum Row in the table
    total_sum = df["total_price"].sum()
    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, border=1)
    pdf.cell(w=50, h=8, border=1)
    pdf.cell(w=30, h=8, border=1)
    pdf.cell(w=30, h=8, border=1)
    pdf.cell(w=30, h=8, txt = str(total_sum), border=1, ln=1)

    # Total sum String
    pdf.set_font(family="Times", size=14, style="BI")
    pdf.cell(w=25, h=8, txt=f"Total Sum: {total_sum}", ln=1)

    # Company name and logo
    pdf.set_font(family="Times", size=14, style="BI")
    pdf.cell(w=25, h=8, txt="PythonHow")
    pdf.image("pythonhow.png", w=10)
    pdf.output(f"PDFs/{filename}.pdf")


